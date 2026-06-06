"""Carga histórica (one-off) desde el Excel de MAGyP.

Lee `descarga_molienda_oleaginosas_historico.xlsx`, toma la sección GRANOS
OLEAGINOSOS (columnas C-I: soja, girasol, lino, maní, algodón, cártamo, canola),
calcula el total y lo inserta con estado=NULL (histórico). Idempotente: usa
insert_if_changed, así que re-correrlo no duplica meses ya cargados.

NO se importa desde main.py.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys

import openpyxl

import db
from magyp import GRANOS

DEFAULT_XLSX = "descarga_molienda_oleaginosas_historico.xlsx"
SHEET = "Molienda Oleaginosas"
FIRST_DATA_ROW = 6          # fila 5 = headers
YEAR_COL, MONTH_COL = 1, 2  # A, B
GRANOS_COLS = range(3, 10)  # C..I = 7 granos en el orden de GRANOS
FUENTE = "excel historico"


def read_rows(path: str) -> list[tuple[dt.date, dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    rows: list[tuple[dt.date, dict]] = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        year = ws.cell(r, YEAR_COL).value
        month = ws.cell(r, MONTH_COL).value
        if year in (None, "") or month in (None, ""):
            continue
        try:
            date = dt.date(int(year), int(month), 1)
        except (ValueError, TypeError):
            continue
        granos = []
        for c in GRANOS_COLS:
            v = ws.cell(r, c).value
            granos.append(float(v) if v not in (None, "") else 0.0)
        row = {g: v for g, v in zip(GRANOS, granos)}
        row["valor"] = float(sum(granos))
        rows.append((date, row))
    wb.close()
    return rows


def main() -> None:
    ap = argparse.ArgumentParser(description="Carga histórica del Excel (one-off).")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help="ruta al Excel")
    ap.add_argument("--force", action="store_true", help="re-insertar aunque no cambie")
    args = ap.parse_args()

    rows = read_rows(args.xlsx)
    if not rows:
        print("No se leyeron filas del Excel.", file=sys.stderr)
        sys.exit(1)
    print(f"Filas leídas: {len(rows)}  rango: {rows[0][0]} .. {rows[-1][0]}")

    conn = db.get_conn()
    inserted = 0
    try:
        for date, row in rows:
            if db.insert_if_changed(conn, date, row, estado=None,
                                    fuente=FUENTE, force=args.force):
                inserted += 1
    finally:
        conn.close()
    print(f"Insertados (nuevos/cambiados): {inserted}  |  sin cambios: {len(rows) - inserted}")


if __name__ == "__main__":
    main()
